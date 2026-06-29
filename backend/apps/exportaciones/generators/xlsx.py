import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


def _estilo_cabecera(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1A56DB')
        cell.alignment = Alignment(horizontal='center')


def exportar_inscritos_xlsx(inscripciones):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inscritos'
    headers = ['Nombre', 'Apellidos', 'DNI', 'Tutor', 'Email Tutor', 'Estado', '% Pagado']
    _estilo_cabecera(ws, headers)
    for ins in inscripciones:
        ws.append([
            ins.alumno.nombre,
            ins.alumno.apellidos,
            ins.alumno.dni,
            ins.padre_tutor.usuario.nombre if ins.padre_tutor else '',
            ins.padre_tutor.usuario.email if ins.padre_tutor else '',
            ins.estado,
            float(ins.porcentaje_pagado),
        ])
    output = io.BytesIO()
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inscritos.xlsx"'
    return response


def exportar_pagos_xlsx(pagos):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pagos'
    headers = ['Alumno', 'Importe', 'Fecha', 'Metodo', 'Estado']
    _estilo_cabecera(ws, headers)
    for pago in pagos:
        ws.append([
            pago.inscripcion.alumno.nombre + ' ' + pago.inscripcion.alumno.apellidos,
            float(pago.importe),
            str(pago.fecha_pago),
            pago.metodo_pago,
            pago.estado,
        ])
    output = io.BytesIO()
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pagos.xlsx"'
    return response


def exportar_documentacion_xlsx(documentos):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Documentacion'
    headers = ['Alumno', 'Documento', 'Estado', 'Archivo', 'Motivo Rechazo']
    _estilo_cabecera(ws, headers)
    for doc in documentos:
        ws.append([
            doc.inscripcion.alumno.nombre + ' ' + doc.inscripcion.alumno.apellidos,
            doc.documento_requerido.nombre,
            doc.estado,
            doc.nombre_archivo,
            doc.motivo_rechazo or '',
        ])
    output = io.BytesIO()
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="documentacion.xlsx"'
    return response